import os
import base64
import pandas as pd
import requests

from azure.identity import ClientSecretCredential
from azure.mgmt.recoveryservicesbackup import RecoveryServicesBackupClient

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


# ==========================
# SERVICE PRINCIPAL #1 (AZURE REPORTING)
# ==========================
AZURE_TENANT_ID = os.getenv("AZURE_TENANT_ID")
AZURE_CLIENT_ID = os.getenv("AZURE_CLIENT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")

credential = ClientSecretCredential(
    tenant_id=AZURE_TENANT_ID,
    client_id=AZURE_CLIENT_ID,
    client_secret=AZURE_CLIENT_SECRET
)


# ==========================
# SERVICE PRINCIPAL #2 (MAIL / GRAPH API)
# ==========================
MAIL_TENANT_ID = os.getenv("MAIL_TENANT_ID")
MAIL_CLIENT_ID = os.getenv("MAIL_CLIENT_ID")
MAIL_CLIENT_SECRET = os.getenv("MAIL_CLIENT_SECRET")

SENDER_EMAIL = os.getenv("SENDER_EMAIL")
RECIPIENT_EMAIL = os.getenv("RECIPIENT_EMAIL")


# ==========================
# VAULT CONFIG
# ==========================
vaults = [
    {
        "subscription_id": "e9e9395f-004e-4e61-99a2-08a0347d5b14",
        "resource_group": "AZE-RG-BACKUPS-NonProduction",
        "vault_name": "AZE-INF-BACKUP-2"
    },
    {
        "subscription_id": "e8328d3b-7c5e-4aa5-b321-eeb887f1fc6b",
        "resource_group": "AZE-RG-BACKUPS",
        "vault_name": "AZE-INF-BACKUP-1"
    },
    {
        "subscription_id": "e8328d3b-7c5e-4aa5-b321-eeb887f1fc6b",
        "resource_group": "AZW-RG-ASR",
        "vault_name": "AZW-RSV-ASR-01"
    }
]

data = []


# ==========================
# FETCH BACKUP REPORT
# ==========================
for v in vaults:
    print(f"\n🔹 Processing Vault: {v['vault_name']}")

    backup_client = RecoveryServicesBackupClient(
        credential,
        v["subscription_id"]
    )

    try:
        items = backup_client.backup_protected_items.list(
            vault_name=v["vault_name"],
            resource_group_name=v["resource_group"]
        )

        for item in items:
            props = item.properties

            # Only VM backups
            if "AzureIaasVM" not in str(props.backup_management_type):
                continue

            source_id = getattr(props, "source_resource_id", "")

            # Extract resource group
            resource_group = source_id.split("/")[4] if source_id else "N/A"

            # Extract VM name
            azure_resource = source_id.split("/")[-1] if source_id else "N/A"

            # Clean Backup Item Name
            backup_item_name = item.name.split(";")[-1] if item.name else "N/A"

            # Latest Recovery Point
            last_rp = getattr(props, "last_recovery_point", None)
            if last_rp:
                last_rp = last_rp.strftime("%Y-%m-%d")
            else:
                last_rp = "N/A"

            # Protection State
            protection_state = getattr(props, "protection_state", "N/A")

            # Resource State
            resource_state = getattr(props, "resource_state", None)

            if not resource_state:
                if not source_id:
                    resource_state = "VM Not Found"
                else:
                    resource_state = "VM Active"

            # Append
            data.append({
                "Backup Item": backup_item_name,
                "Resource Group": resource_group,
                "Protection State": protection_state,
                "Health Check Status": getattr(props, "health_status", "N/A"),
                "Latest Recovery Point": last_rp,
                "Resource State": resource_state,
                "Azure Resource": azure_resource,
                "Vault": v["vault_name"]
            })

    except Exception as e:
        print(f"❌ Error in {v['vault_name']}: {e}")


# ==========================
# CREATE DATAFRAME
# ==========================
df = pd.DataFrame(data)

if df.empty:
    print("❌ No backup data found. Check Azure permissions.")
    exit()

df = df[
    [
        "Backup Item",
        "Resource Group",
        "Protection State",
        "Health Check Status",
        "Latest Recovery Point",
        "Resource State",
        "Azure Resource",
        "Vault"
    ]
]


# ==========================
# EXPORT EXCEL
# ==========================
file_name = "Backup_Explorer_Report.xlsx"
df.to_excel(file_name, index=False)


# ==========================
# STYLE EXCEL
# ==========================
wb = load_workbook(file_name)
ws = wb.active

header_fill = PatternFill(
    start_color="87CEEB",
    end_color="87CEEB",
    fill_type="solid"
)

header_font = Font(
    bold=True,
    color="000000"
)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

center_align = Alignment(
    horizontal="center",
    vertical="center"
)

# Header
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align

# Data
for row in ws.iter_rows(
    min_row=2,
    max_row=ws.max_row,
    max_col=ws.max_column
):
    for cell in row:
        cell.border = thin_border
        cell.alignment = center_align

# Auto width
for column_cells in ws.columns:
    max_length = 0
    column_letter = column_cells[0].column_letter

    for cell in column_cells:
        try:
            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
        except Exception:
            pass

    ws.column_dimensions[column_letter].width = max_length + 2

wb.save(file_name)

print(f"\n✅ Report Generated: {file_name}")


# ==========================
# BUILD MAIL SUMMARY
# ==========================
total_backup_items = len(df)

successful_backup_items = len(
    df[
        (df["Health Check Status"].str.lower() == "passed") &
        (df["Protection State"].str.lower() == "protected")
    ]
)

non_protected_items = df[
    df["Protection State"].str.lower() != "protected"
]

failed_backup_items = len(non_protected_items)

non_protected_servers = non_protected_items["Azure Resource"].tolist()

if non_protected_servers:
    server_list = ", ".join(non_protected_servers)
    note_section = f"""
    <p><b>Note:</b></p>
    <p>
    The following servers — {server_list} —
    backups are in non-protected state.
    </p>
    """
else:
    note_section = """
    <p><b>Note:</b></p>
    <p>All backup items are currently protected.</p>
    """


# ==========================
# EMAIL BODY
# ==========================
email_html_body = f"""
<p>Hi Sam,</p>

<p>
Please find the attached report for the Weekly Backup.
Below is the status of all backup items.
</p>

<table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse;">
    <tr style="background-color:#1F4E78; color:white;">
        <th>Backup Status</th>
        <th>Count</th>
    </tr>
    <tr>
        <td><b>Total Backup Items</b></td>
        <td>{total_backup_items}</td>
    </tr>
    <tr>
        <td><b>Successful Backup Items</b></td>
        <td>{successful_backup_items}</td>
    </tr>
    <tr>
        <td><b>Failed / Non-Protected Items</b></td>
        <td>{failed_backup_items}</td>
    </tr>
</table>

<br>

{note_section}

<p>Regards,<br>Automation System</p>
"""


# ==========================
# GRAPH TOKEN FUNCTION
# ==========================
def get_graph_token():
    url = f"https://login.microsoftonline.com/{MAIL_TENANT_ID}/oauth2/v2.0/token"

    payload = {
        "client_id": MAIL_CLIENT_ID,
        "client_secret": MAIL_CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials"
    }

    response = requests.post(url, data=payload)
    response.raise_for_status()

    return response.json()["access_token"]


# ==========================
# SEND EMAIL FUNCTION
# ==========================
def send_email_with_attachment():
    token = get_graph_token()

    with open(file_name, "rb") as f:
        encoded_file = base64.b64encode(
            f.read()
        ).decode("utf-8")

    email_body = {
        "message": {
            "subject": "Weekly Backup Explorer Report",
            "body": {
                "contentType": "HTML",
                "content": email_html_body
            },
            "toRecipients": [
                {
                    "emailAddress": {
                        "address": RECIPIENT_EMAIL
                    }
                }
            ],
            "attachments": [
                {
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": file_name,
                    "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "contentBytes": encoded_file
                }
            ]
        },
        "saveToSentItems": "true"
    }

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    endpoint = f"https://graph.microsoft.com/v1.0/users/{SENDER_EMAIL}/sendMail"

    response = requests.post(
        endpoint,
        headers=headers,
        json=email_body
    )

    if response.status_code == 202:
        print("✅ Email sent successfully.")
    else:
        print(f"❌ Email failed: {response.status_code}")
        print(response.text)


# ==========================
# SEND REPORT
# ==========================
send_email_with_attachment()
